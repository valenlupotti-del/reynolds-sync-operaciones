from __future__ import annotations
import io
import os
import json
import uuid
import smtplib
import logging
import threading
import time
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, request, jsonify
import requests
import sync_properties
import sync_contacts
import create_campaigns
import tasacion as tasacion_module

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__)

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

GMAIL_USER = os.environ.get("GMAIL_USER", "valenlupotti@gmail.com")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "bnsa axel vpqq hvhr")
GMAIL_FROM_NAME = os.environ.get("GMAIL_FROM_NAME", "Reynolds Propiedades")

TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")

ROLE_TEMPLATE = {
    "comprador":   "email_comprador.html",
    "vendedor":    "email_vendedor.html",
    "inquilino":   "email_inquilino.html",
    "propietario": "email_propietario.html",
    "tasacion":    "email_tasacion.html",
    "tasación":    "email_tasacion.html",
}

TALLY_FIELD_MAP = {
    "Asesor":                           "asesor_nombre",
    "Sucursal":                         "sucursal",
    "Fecha de firma":                   "fecha_firma",
    "Dirección de la propiedad":        "direccion",
    "Tipo de propiedad":                "tipo_propiedad",
    "Tipo de Operación":                "tipo_operacion",
    "Monto en USD":                     "monto",
    "Porcentaje de comisión":           "porcentaje_comision",
    "Monto mensual":                    "monto_mensual",
    "Plazo en meses?":                  "plazo",
    "Valor Tasación":                   "comision",
    "Nombre del cliente":               "cliente_nombre",
    "Email del cliente":                "cliente_email",
    "Rol del cliente":                  "cliente_rol",
    "Nombre de la contraparte":         "contraparte_nombre",
    "Email de la contraparte":          "contraparte_email",
    "Rol de la contraparte":            "contraparte_rol",
    "Observaciones":                    "observaciones",
}

TALLY_TASACION_MAP = {
    "Nombre del asesor":           "asesor_nombre",
    "Email del asesor":            "asesor_email",
    "Nombre del cliente":          "cliente_nombre",
    "Email del cliente":           "cliente_email",
    "Dirección de la propiedad":   "direccion",
    "Barrio":                      "barrio",
    "Tipo de propiedad":           "tipo_propiedad",
    "Superficie (m²)":             "superficie",
    "Precio estimado (USD)":       "precio_estimado",
    "Link comparable 1":           "link_1",
    "Link comparable 2":           "link_2",
    "Link comparable 3":           "link_3",
    "Link comparable 4":           "link_4",
    "Observaciones del asesor":    "observaciones",
}

# In-memory store for pending asesor approvals: token -> {html, cliente_email, subject, expires}
_pending_tasaciones: dict[str, dict] = {}


def load_template(role: str) -> str | None:
    filename = ROLE_TEMPLATE.get(role.lower().strip())
    if not filename:
        return None
    path = os.path.join(TEMPLATES_DIR, filename)
    try:
        with open(path, encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        log.error("Template not found: %s", path)
        return None


def send_email(to_email: str, subject: str, html_body: str, attachments: list[tuple[str, bytes]] | None = None):
    """
    Send an HTML email with optional attachments.
    attachments: list of (filename, bytes) tuples
    """
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = f"{GMAIL_FROM_NAME} <{GMAIL_USER}>"
    msg["To"] = to_email

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    for filename, data in (attachments or []):
        part = MIMEBase("application", "octet-stream")
        part.set_payload(data)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, to_email, msg.as_string())
    log.info("Email sent to %s", to_email)


def build_tasacion_excel(subject_prop: dict, comparables: list[dict | None], ai_opinion: str = "") -> bytes:
    """Generate an Excel comparison workbook and return as bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed")
        return b""

    wb = Workbook()

    # ---- Sheet 1: Comparativa ----
    ws = wb.active
    ws.title = "Comparativa"

    AZUL  = "1B3A6B"
    ROJO  = "C0392B"
    VERDE = "16A34A"
    GRIS  = "F3F4F6"
    AMARILLO = "FEF9C3"

    def cell_style(cell, bold=False, bg=None, font_color="000000", size=11, align="left", wrap=False):
        cell.font = Font(bold=bold, color=font_color, size=size)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "INFORME COMPARATIVO DE TASACION — Reynolds Propiedades"
    cell_style(ws["A1"], bold=True, bg=AZUL, font_color="FFFFFF", size=13, align="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell_style(ws["A2"], bg="E5E7EB", font_color="6B7280", size=9, align="center")
    ws.row_dimensions[2].height = 16

    # Header row
    headers = ["Campo", "Tu propiedad"] + [f"Comparable {i+1}" for i in range(len([c for c in comparables if c is not None]))]
    ws.row_dimensions[4].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        bg = AMARILLO if col == 2 else AZUL
        fc = "000000" if col == 2 else "FFFFFF"
        cell_style(c, bold=True, bg=bg, font_color=fc, size=11, align="center")
        c.border = border

    # Data rows
    valid_comps = [c for c in comparables if c is not None]
    fields = [
        ("Barrio",       lambda p: p.get("barrio") or "—"),
        ("Direccion",    lambda p: (p.get("direccion") or "—")[:40]),
        ("Tipo",         lambda p: p.get("tipo") or "—"),
        ("Superficie",   lambda p: f"{int(p['superficie'])} m²" if p.get("superficie") else "—"),
        ("Precio (USD)", lambda p: f"USD {int(p['precio']):,}".replace(",", ".") if p.get("precio") else "—"),
        ("USD/m²",       lambda p: f"USD {int(p['usd_m2']):,}/m²".replace(",", ".") if p.get("usd_m2") else "—"),
        ("Ambientes",    lambda p: str(p.get("ambientes") or "—")),
    ]

    for row_idx, (label, getter) in enumerate(fields, 5):
        ws.row_dimensions[row_idx].height = 20
        c = ws.cell(row=row_idx, column=1, value=label)
        cell_style(c, bold=True, bg=GRIS, size=10)
        c.border = border

        # Subject property
        c2 = ws.cell(row=row_idx, column=2, value=getter(subject_prop))
        cell_style(c2, bg=AMARILLO, size=10, align="center")
        c2.border = border

        for ci, comp in enumerate(valid_comps, 3):
            val = getter(comp)
            c3 = ws.cell(row=row_idx, column=ci, value=val)
            cell_style(c3, size=10, align="center")
            c3.border = border

    # USD/m2 average row
    usd_m2_vals = [c["usd_m2"] for c in valid_comps if c.get("usd_m2")]
    if usd_m2_vals:
        avg = round(sum(usd_m2_vals) / len(usd_m2_vals))
        row = len(fields) + 6
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = f"Promedio USD/m² comparables:  USD {avg:,}/m²".replace(",", ".")
        cell_style(ws[f"A{row}"], bold=True, bg=VERDE, font_color="FFFFFF", size=11)
        ws.row_dimensions[row].height = 22

        sup = subject_prop.get("superficie")
        if sup:
            row2 = row + 1
            val_est = round(avg * sup)
            ws.merge_cells(f"A{row2}:B{row2}")
            ws[f"A{row2}"] = f"Valor estimado ({int(sup)} m²):  USD {val_est:,}".replace(",", ".")
            cell_style(ws[f"A{row2}"], bold=True, bg=ROJO, font_color="FFFFFF", size=11)
            ws.row_dimensions[row2].height = 22

    # Column widths
    col_widths = [20, 22] + [22] * len(valid_comps)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: Opinion IA ----
    if ai_opinion:
        ws2 = wb.create_sheet("Opinion IA")
        ws2.merge_cells("A1:D1")
        ws2["A1"] = "Opinion del mercado — generada por IA"
        cell_style(ws2["A1"], bold=True, bg=AZUL, font_color="FFFFFF", size=12, align="center")
        ws2.row_dimensions[1].height = 26

        ws2.merge_cells("A3:D20")
        ws2["A3"] = ai_opinion
        ws2["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws2["A3"].font = Font(size=11)
        ws2.row_dimensions[3].height = 200
        for col in ["A", "B", "C", "D"]:
            ws2.column_dimensions[col].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_tasacion_opinion(subject_prop: dict, comparables: list[dict | None], avg_usd_m2: int | None) -> str:
    """Use OpenAI to generate a professional valuation opinion paragraph."""
    openai_key = os.environ.get("OPENAI_KEY", "")
    if not openai_key:
        return ""

    comp_lines = []
    for i, c in enumerate(comparables, 1):
        if c:
            sup = f"{int(c['superficie'])} m²" if c.get("superficie") else "?"
            precio = f"USD {int(c['precio']):,}".replace(",", ".") if c.get("precio") else "?"
            usd_m2 = f"USD {int(c['usd_m2']):,}/m²".replace(",", ".") if c.get("usd_m2") else "?"
            comp_lines.append(f"  Comp {i}: {c.get('barrio','?')} — {c.get('tipo','?')} — {sup} — {precio} ({usd_m2})")

    subject_lines = [
        f"Barrio: {subject_prop.get('barrio','?')}",
        f"Tipo: {subject_prop.get('tipo_propiedad') or subject_prop.get('tipo','?')}",
        f"Superficie: {int(subject_prop['superficie'])} m²" if subject_prop.get("superficie") else "",
        f"Precio estimado por el asesor: USD {int(subject_prop['precio']):,}".replace(",", ".") if subject_prop.get("precio") else "",
        f"Promedio USD/m² de comparables: USD {avg_usd_m2:,}/m²".replace(",", ".") if avg_usd_m2 else "",
    ]

    prompt = f"""Sos un analista inmobiliario senior de Reynolds Propiedades, una inmobiliaria premium de zona norte de Buenos Aires.

Propiedades comparable analizadas:
{chr(10).join(comp_lines)}

Propiedad a tasar:
{chr(10).join(l for l in subject_lines if l)}

Escribi un parrafo profesional (4-6 oraciones) que:
1. Describa el posicionamiento de precio de la propiedad respecto al mercado
2. Mencione si el precio esta por encima, dentro o por debajo del rango de comparables
3. De una opinion sobre el valor de mercado
4. Sea en tono profesional pero accesible, sin tecnicismos excesivos
5. Sea en castellano argentino

Respondé SOLO con el parrafo, sin titulos ni explicaciones."""

    try:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {openai_key}", "Content-Type": "application/json"},
            json={
                "model": "gpt-4o-mini",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7,
                "max_tokens": 300,
            },
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        log.error("OpenAI opinion error: %s", e)
        return ""


def parse_tally_payload(payload: dict) -> dict:
    data = {}
    fields = payload.get("data", {}).get("fields", [])
    log.info("Tally fields received: %s", [f.get("label") for f in fields])
    for field in fields:
        label = field.get("label", "")
        value = field.get("value")
        options = field.get("options", [])

        # Resolve UUID option IDs to text labels
        if options and value is not None:
            option_map = {o["id"]: o["text"] for o in options if "id" in o and "text" in o}
            if isinstance(value, list):
                value = [option_map.get(v, v) for v in value]
            else:
                value = option_map.get(value, value)

        # Unwrap single-item lists
        if isinstance(value, list) and len(value) == 1:
            value = value[0]

        key = TALLY_FIELD_MAP.get(label)
        if key:
            data[key] = value
    return data


def insert_operacion(data: dict) -> dict | None:
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    row = {k: v for k, v in data.items() if v not in (None, "", [])}
    row["fecha_carga"] = datetime.utcnow().isoformat()

    # Convert numeric strings
    for field in ("monto", "monto_mensual", "plazo", "porcentaje_comision", "comision"):
        if field in row:
            try:
                row[field] = float(str(row[field]).replace(",", "."))
            except (ValueError, TypeError):
                row.pop(field, None)

    # Calculate comision
    tipo = (row.get("tipo_operacion") or "").lower()
    pct = row.get("porcentaje_comision")
    if pct and "alquiler" in tipo:
        mensual = row.get("monto_mensual")
        plazo = row.get("plazo")
        if mensual and plazo:
            row["comision"] = round(mensual * plazo * pct / 100, 2)
    elif pct and "venta" in tipo:
        monto = row.get("monto")
        if monto:
            row["comision"] = round(monto * pct / 100, 2)

    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/operaciones",
        headers=headers,
        json=row,
        timeout=15,
    )
    if not resp.ok:
        log.error("Supabase error %s: %s", resp.status_code, resp.text)
    resp.raise_for_status()
    result = resp.json()
    return result[0] if result else None


def notify_client(name: str, email: str, role: str, asesor: str):
    template = load_template(role)
    if not template:
        log.warning("No template for role '%s', skipping email to %s", role, email)
        return
    html = template.replace("{NOMBRE_CLIENTE}", name).replace("{ASESOR}", asesor)

    role_subjects = {
        "comprador":   "¡Tu nueva propiedad ya es tuya! 🏡",
        "vendedor":    "¡Tu propiedad encontró su dueño! ✨",
        "inquilino":   "¡Tu nuevo hogar te está esperando! 🔑",
        "propietario": "Tu propiedad ya tiene nuevos inquilinos 🤝",
        "tasacion":    "¡Tu tasación está lista! 📊",
        "tasación":    "¡Tu tasación está lista! 📊",
    }
    subject = role_subjects.get(role.lower().strip(), "Reynolds Propiedades — gracias por confiar en nosotros")
    send_email(email, subject, html)


@app.route("/webhook/operaciones", methods=["POST"])
def webhook():
    payload = request.get_json(force=True, silent=True)
    if not payload:
        return jsonify({"error": "invalid payload"}), 400

    log.info("Tally submission received")

    try:
        data = parse_tally_payload(payload)
    except Exception as e:
        log.exception("Error parsing Tally payload")
        return jsonify({"error": str(e)}), 422

    if not data:
        log.warning("Empty data after parsing, raw payload: %s", json.dumps(payload)[:500])
        return jsonify({"error": "no recognizable fields"}), 422

    # Insert into Supabase
    try:
        insert_operacion(data)
        log.info("Operation inserted for %s", data.get("cliente_nombre", "?"))
    except Exception as e:
        log.exception("Supabase insert failed")
        return jsonify({"error": f"DB insert failed: {e}"}), 500

    asesor = data.get("asesor_nombre", "tu asesor")

    # Send email to main client
    cliente_email = data.get("cliente_email", "")
    cliente_nombre = data.get("cliente_nombre", "")
    cliente_rol = data.get("cliente_rol", "")
    if cliente_email and cliente_nombre and cliente_rol:
        try:
            notify_client(cliente_nombre, cliente_email, cliente_rol, asesor)
        except Exception as e:
            log.error("Failed to email client %s: %s", cliente_email, e)

    # Send email to counterpart if present
    contra_email = data.get("contraparte_email", "")
    contra_nombre = data.get("contraparte_nombre", "")
    contra_rol = data.get("contraparte_rol", "")
    if contra_email and contra_nombre and contra_rol:
        try:
            notify_client(contra_nombre, contra_email, contra_rol, asesor)
        except Exception as e:
            log.error("Failed to email counterpart %s: %s", contra_email, e)

    return jsonify({"ok": True}), 200


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"}), 200


@app.route("/sync/contacts", methods=["GET", "POST"])
def trigger_contacts_sync():
    def _run():
        try:
            log.info("Manual contacts sync triggered via endpoint")
            sync_contacts.run()
            log.info("Manual contacts sync complete")
        except Exception:
            log.exception("Error in manual contacts sync")
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "sync started"}), 200


# ---------------------------------------------------------------------------
# Tasacion webhook — asesor fills Tally with 4 comparable links
# ---------------------------------------------------------------------------

def _parse_tally_tasacion(payload: dict) -> dict:
    data = {}
    fields = payload.get("data", {}).get("fields", [])
    for field in fields:
        label = field.get("label", "")
        value = field.get("value")
        options = field.get("options", [])
        if options and value is not None:
            option_map = {o["id"]: o["text"] for o in options if "id" in o and "text" in o}
            if isinstance(value, list):
                value = [option_map.get(v, v) for v in value]
            else:
                value = option_map.get(value, value)
        if isinstance(value, list) and len(value) == 1:
            value = value[0]
        key = TALLY_TASACION_MAP.get(label)
        if key:
            data[key] = value
    return data


def _comparable_row(idx: int, prop: dict | None, is_subject: bool = False) -> str:
    if prop is None:
        return f"""<td style="padding:14px 12px;border-right:1px solid #e5e7eb;text-align:center;color:#9ca3af;font-size:13px;">
            No disponible
        </td>"""

    bg = "#fef3c7" if is_subject else "#ffffff"
    label = "<strong>Tu propiedad</strong>" if is_subject else f"Comp. {idx}"
    barrio = prop.get("barrio") or "—"
    direccion = (prop.get("direccion") or "—")[:40]
    tipo = prop.get("tipo") or "—"
    sup = f"{int(prop['superficie'])} m²" if prop.get("superficie") else "—"
    precio = f"USD {int(prop['precio']):,}".replace(",", ".") if prop.get("precio") else "—"
    usd_m2 = f"USD {int(prop['usd_m2']):,}/m²".replace(",", ".") if prop.get("usd_m2") else "—"
    url = prop.get("url", "#")

    return f"""<td style="padding:14px 12px;border-right:1px solid #e5e7eb;background:{bg};vertical-align:top;">
        <p style="margin:0 0 2px;font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:1px;">{label}</p>
        <p style="margin:0 0 6px;font-size:13px;font-weight:700;color:#111827;">{barrio}</p>
        <p style="margin:0 0 8px;font-size:11px;color:#6b7280;">{tipo} · {direccion}</p>
        <table cellpadding="0" cellspacing="0" width="100%">
          <tr><td style="font-size:11px;color:#6b7280;padding-bottom:3px;">Superficie</td><td style="font-size:12px;font-weight:600;color:#374151;text-align:right;">{sup}</td></tr>
          <tr><td style="font-size:11px;color:#6b7280;padding-bottom:3px;">Precio</td><td style="font-size:13px;font-weight:700;color:#1d4ed8;text-align:right;">{precio}</td></tr>
          <tr><td style="font-size:11px;color:#6b7280;">USD/m²</td><td style="font-size:12px;font-weight:600;color:#059669;text-align:right;">{usd_m2}</td></tr>
        </table>
        {"" if is_subject else f'<p style="margin:8px 0 0;"><a href="{url}" style="font-size:10px;color:#2563eb;">Ver en Argenprop →</a></p>'}
    </td>"""


def _build_tasacion_asesor_html(
    subject_prop: dict,
    comparables: list[dict | None],
    asesor_nombre: str,
    cliente_nombre: str,
    cliente_email: str,
    observaciones: str,
    approve_url: str,
    ai_opinion: str = "",
) -> str:
    # Compute average USD/m2 from valid comparables
    valid_usd_m2 = [c["usd_m2"] for c in comparables if c and c.get("usd_m2")]
    avg_usd_m2 = round(sum(valid_usd_m2) / len(valid_usd_m2)) if valid_usd_m2 else None

    sup = subject_prop.get("superficie")
    estimado_block = ""
    if avg_usd_m2 and sup:
        val_estimado = round(avg_usd_m2 * sup)
        estimado_block = f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin:24px 0;border-radius:8px;background:#ecfdf5;border:1px solid #6ee7b7;">
          <tr><td style="padding:20px 24px;">
            <p style="margin:0 0 4px;font-size:11px;font-weight:700;color:#065f46;letter-spacing:2px;text-transform:uppercase;">Estimacion automatica</p>
            <p style="margin:0;font-size:22px;font-weight:800;color:#065f46;">USD {val_estimado:,} <span style="font-size:14px;font-weight:400;">para {int(sup)} m²</span></p>
            <p style="margin:4px 0 0;font-size:12px;color:#047857;">Basado en promedio USD {avg_usd_m2:,}/m² de los {len(valid_usd_m2)} comparables</p>
          </td></tr>
        </table>""".replace(",", ".")

    # Build comparison table columns
    subject_col = _comparable_row(0, subject_prop, is_subject=True)
    comp_cols = "".join(_comparable_row(i + 1, c) for i, c in enumerate(comparables))

    obs_block = ""
    if observaciones:
        obs_block = f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin:20px 0;border-radius:8px;background:#fffbeb;border:1px solid #fcd34d;">
          <tr><td style="padding:16px 20px;">
            <p style="margin:0 0 4px;font-size:11px;font-weight:700;color:#92400e;letter-spacing:2px;text-transform:uppercase;">Notas del asesor</p>
            <p style="margin:0;font-size:13px;color:#78350f;">{observaciones}</p>
          </td></tr>
        </table>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;">
  <tr><td align="center" style="padding:32px 10px;">
    <table width="620" cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;">

      <tr><td style="padding-bottom:20px;text-align:center;">
        <img src="https://d1v2p1s05qqabi.cloudfront.net/sites/38/media/163287410792.jpeg?v=18" alt="Reynolds Propiedades" width="110" style="display:block;margin:0 auto;border-radius:6px;">
      </td></tr>

      <tr><td style="background:#1B3A6B;border-radius:12px 12px 0 0;padding:36px 40px 28px;">
        <p style="margin:0 0 6px;font-size:11px;font-weight:700;color:#93c5fd;letter-spacing:3px;text-transform:uppercase;">Revision previa al envio</p>
        <h1 style="margin:0 0 10px;font-size:26px;font-weight:800;color:#ffffff;">Tasacion lista para revisar</h1>
        <p style="margin:0;font-size:14px;color:#bfdbfe;">Cliente: <strong style="color:#ffffff;">{cliente_nombre}</strong> ({cliente_email})</p>
      </td></tr>

      <tr><td style="background:#ffffff;padding:32px 40px;">

        <p style="margin:0 0 20px;font-size:14px;color:#374151;">Hola <strong>{asesor_nombre}</strong>, el sistema armo la comparativa con los links que cargaste. Revisa los datos y si todo esta bien, aproba el envio al cliente.</p>

        <!-- Comparison table -->
        <div style="overflow-x:auto;">
        <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
          <tr style="background:#f9fafb;">
            {subject_col}
            {comp_cols}
          </tr>
        </table>
        </div>

        {estimado_block}
        {obs_block}
        {"" if not ai_opinion else f'''
        <table width="100%" cellpadding="0" cellspacing="0" style="margin:20px 0;border-radius:8px;background:#f0f4ff;border:1px solid #c7d2fe;">
          <tr><td style="padding:18px 20px;">
            <p style="margin:0 0 6px;font-size:11px;font-weight:700;color:#3730a3;letter-spacing:2px;text-transform:uppercase;">Opinion IA del mercado</p>
            <p style="margin:0;font-size:13px;color:#1e1b4b;line-height:1.7;">{ai_opinion}</p>
          </td></tr>
        </table>'''}
        <p style="margin:0 0 8px;font-size:12px;color:#9ca3af;">El Excel con la comparativa completa va adjunto a este email.</p>

        <!-- Approve button -->
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:24px;">
          <tr><td style="text-align:center;">
            <p style="margin:0 0 8px;font-size:13px;color:#6b7280;">Si los datos son correctos, hace clic para enviar al cliente:</p>
            <a href="{approve_url}" style="display:inline-block;padding:16px 40px;background:#16a34a;color:#ffffff;font-size:15px;font-weight:700;text-decoration:none;border-radius:8px;">Aprobar y enviar al cliente</a>
            <p style="margin:12px 0 0;font-size:11px;color:#9ca3af;">Este link expira en 48 horas. Si no queres enviar nada, ignora este email.</p>
          </td></tr>
        </table>

      </td></tr>

      <tr><td style="padding:20px 0;text-align:center;">
        <p style="margin:0;font-size:11px;color:#9ca3af;">Reynolds Propiedades &middot; Sistema interno</p>
      </td></tr>

    </table>
  </td></tr>
</table>
</body></html>"""


def _build_tasacion_cliente_html(
    subject_prop: dict,
    comparables: list[dict | None],
    cliente_nombre: str,
    asesor_nombre: str,
    observaciones: str,
    ai_opinion: str = "",
) -> str:
    valid_usd_m2 = [c["usd_m2"] for c in comparables if c and c.get("usd_m2")]
    avg_usd_m2 = round(sum(valid_usd_m2) / len(valid_usd_m2)) if valid_usd_m2 else None
    sup = subject_prop.get("superficie")

    estimado_block = ""
    if avg_usd_m2 and sup:
        val = round(avg_usd_m2 * sup)
        estimado_block = f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin:24px 0;border-radius:10px;background:#f0fdf4;border:2px solid #86efac;">
          <tr><td style="padding:24px 28px;text-align:center;">
            <p style="margin:0 0 6px;font-size:11px;font-weight:700;color:#166534;letter-spacing:2px;text-transform:uppercase;">Valor estimado de tu propiedad</p>
            <p style="margin:0;font-size:32px;font-weight:800;color:#15803d;">USD {val:,}</p>
            <p style="margin:6px 0 0;font-size:12px;color:#166534;">Basado en {len(valid_usd_m2)} propiedades comparables · Precio mediano USD {avg_usd_m2:,}/m²</p>
          </td></tr>
        </table>""".replace(",", ".")

    subject_col = _comparable_row(0, subject_prop, is_subject=True)
    comp_cols = "".join(_comparable_row(i + 1, c) for i, c in enumerate(comparables))

    obs_block = ""
    if observaciones:
        obs_block = f"""
        <table width="100%" cellpadding="0" cellspacing="0" style="margin:20px 0;border-radius:8px;background:#fffbeb;border:1px solid #fcd34d;">
          <tr><td style="padding:16px 20px;">
            <p style="margin:0 0 4px;font-size:11px;font-weight:700;color:#92400e;letter-spacing:1px;text-transform:uppercase;">Comentario del asesor</p>
            <p style="margin:0;font-size:14px;color:#78350f;line-height:1.6;">{observaciones}</p>
          </td></tr>
        </table>"""

    return f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#ECEEF2;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#ECEEF2;">
  <tr><td align="center" style="padding:32px 10px;">
    <table width="620" cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;">

      <tr><td style="padding-bottom:24px;text-align:center;">
        <img src="https://d1v2p1s05qqabi.cloudfront.net/sites/38/media/163287410792.jpeg?v=18" alt="Reynolds Propiedades" width="130" style="display:block;margin:0 auto;border-radius:6px;">
      </td></tr>

      <tr><td style="background:linear-gradient(135deg,#1B3A6B 0%,#0f2347 100%);border-radius:12px 12px 0 0;padding:52px 40px 40px;text-align:center;">
        <p style="margin:0 0 12px;font-size:12px;font-weight:700;color:#93c5fd;letter-spacing:3px;text-transform:uppercase;">Informe de Tasacion</p>
        <h1 style="margin:0 0 14px;font-size:28px;font-weight:800;color:#ffffff;line-height:1.3;">El valor de tu propiedad,<br>respaldado por el mercado</h1>
        <p style="margin:0 auto;font-size:14px;color:#bfdbfe;max-width:400px;line-height:1.7;">Analizamos propiedades comparables para darte una referencia precisa y objetiva.</p>
      </td></tr>

      <tr><td style="background:#ffffff;border-radius:0 0 12px 12px;padding:36px 40px 44px;">

        <p style="margin:0 0 24px;font-size:15px;color:#374151;line-height:1.7;">Hola <strong>{cliente_nombre}</strong>, preparamos este informe comparativo para que tengas una vision clara del valor de mercado de tu propiedad.</p>

        <!-- Comparison table -->
        <div style="overflow-x:auto;">
        <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
          <tr style="background:#f9fafb;">
            {subject_col}
            {comp_cols}
          </tr>
        </table>
        </div>

        {estimado_block}
        {obs_block}
        {"" if not ai_opinion else f'''
        <table width="100%" cellpadding="0" cellspacing="0" style="margin:20px 0;border-radius:10px;background:#f8faff;border:1px solid #dbeafe;">
          <tr><td style="padding:20px 24px;">
            <p style="margin:0 0 8px;font-size:11px;font-weight:700;color:#1d4ed8;letter-spacing:2px;text-transform:uppercase;">Analisis del mercado</p>
            <p style="margin:0;font-size:14px;color:#1e3a5f;line-height:1.8;">{ai_opinion}</p>
          </td></tr>
        </table>'''}
        <p style="margin:16px 0 0;font-size:12px;color:#9ca3af;">El informe completo con la comparativa detallada va adjunto como archivo Excel.</p>

        <table width="100%" cellpadding="0" cellspacing="0" style="margin:28px 0 0;border-top:1px solid #e5e7eb;padding-top:24px;">
          <tr><td>
            <p style="margin:0 0 4px;font-size:14px;color:#374151;">Cualquier consulta, <strong>{asesor_nombre}</strong> esta a tu disposicion.</p>
            <p style="margin:0;font-size:13px;color:#6b7280;">Equipo Reynolds Propiedades</p>
          </td></tr>
        </table>

      </td></tr>

      <tr><td style="padding:24px 0 0;text-align:center;">
        <p style="margin:0 0 4px;font-size:12px;color:#9ca3af;">Reynolds Propiedades &middot; <a href="https://reynoldspropiedades.com.ar" style="color:#C0392B;text-decoration:none;">reynoldspropiedades.com.ar</a></p>
        <p style="margin:0;font-size:11px;color:#bbbbbb;">Informe generado automaticamente con datos de Argenprop.</p>
      </td></tr>

    </table>
  </td></tr>
</table>
</body></html>"""


@app.route("/webhook/tasacion", methods=["POST"])
def webhook_tasacion():
    """
    Tally webhook for tasacion comparativa.
    Asesor submits 4 Argenprop links + property details.
    System scrapes them, builds comparison, emails asesor for approval.
    """
    payload = request.get_json(force=True, silent=True)
    if not payload:
        return jsonify({"error": "invalid payload"}), 400

    try:
        data = _parse_tally_tasacion(payload)
    except Exception as e:
        log.exception("Error parsing tasacion payload")
        return jsonify({"error": str(e)}), 422

    asesor_email  = data.get("asesor_email", "").strip()
    asesor_nombre = data.get("asesor_nombre", "el asesor").strip()
    cliente_email  = data.get("cliente_email", "").strip()
    cliente_nombre = data.get("cliente_nombre", "el cliente").strip()

    if not asesor_email or not cliente_email:
        return jsonify({"error": "asesor_email and cliente_email are required"}), 422

    # Subject property (from form data — no scraping needed)
    superficie_raw = data.get("superficie")
    superficie = None
    if superficie_raw:
        try:
            superficie = float(str(superficie_raw).replace(",", "."))
        except (ValueError, TypeError):
            pass

    precio_est_raw = data.get("precio_estimado")
    precio_estimado = None
    if precio_est_raw:
        try:
            precio_estimado = float(str(precio_est_raw).replace(".", "").replace(",", ""))
        except (ValueError, TypeError):
            pass

    subject_prop = {
        "barrio":     data.get("barrio") or "—",
        "direccion":  data.get("direccion") or "—",
        "tipo":       data.get("tipo_propiedad") or "Propiedad",
        "superficie": superficie,
        "precio":     precio_estimado,
        "usd_m2":     round(precio_estimado / superficie) if precio_estimado and superficie else None,
        "url":        None,
    }

    # Scrape the 4 comparable links
    links = [data.get(f"link_{i}") for i in range(1, 5)]
    comparables: list[dict | None] = []
    for link in links:
        if not link:
            comparables.append(None)
            continue
        try:
            prop = tasacion_module.scrape_listing_url(link.strip())
            comparables.append(prop)
            log.info("Scraped %s -> %s", link, prop)
        except Exception as e:
            log.error("Failed scraping %s: %s", link, e)
            comparables.append(None)

    # Trim trailing None to only include filled comparables
    while comparables and comparables[-1] is None:
        comparables.pop()

    if not any(c for c in comparables if c):
        return jsonify({"error": "No se pudo scrapelar ningun comparable"}), 422

    observaciones = data.get("observaciones") or ""

    # Compute avg USD/m2 from comparables
    valid_usd_m2 = [c["usd_m2"] for c in comparables if c and c.get("usd_m2")]
    avg_usd_m2 = round(sum(valid_usd_m2) / len(valid_usd_m2)) if valid_usd_m2 else None

    # Generate AI opinion
    log.info("Generating AI opinion for tasacion...")
    ai_opinion = generate_tasacion_opinion(subject_prop, comparables, avg_usd_m2)
    if ai_opinion:
        log.info("AI opinion generated (%d chars)", len(ai_opinion))

    # Generate Excel
    excel_bytes = build_tasacion_excel(subject_prop, comparables, ai_opinion)
    excel_filename = f"tasacion_{tasacion_module.slugify(subject_prop.get('barrio','propiedad'))}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    attachments = [(excel_filename, excel_bytes)] if excel_bytes else []

    # Generate approve token and store pending email
    token = str(uuid.uuid4())
    client_html = _build_tasacion_cliente_html(
        subject_prop, comparables, cliente_nombre, asesor_nombre, observaciones, ai_opinion
    )
    client_subject = f"Tasacion de tu {subject_prop['tipo']} en {subject_prop['barrio']} — Reynolds Propiedades"

    _pending_tasaciones[token] = {
        "html":            client_html,
        "subject":         client_subject,
        "cliente_email":   cliente_email,
        "cliente_nombre":  cliente_nombre,
        "asesor_nombre":   asesor_nombre,
        "attachments":     attachments,
        "created_at":      time.time(),
    }

    # Build the approval URL
    server_url = os.environ.get("SERVER_URL", request.host_url.rstrip("/"))
    approve_url = f"{server_url}/tasacion/aprobar?token={token}"

    # Build and send asesor review email (with Excel attached)
    asesor_html = _build_tasacion_asesor_html(
        subject_prop, comparables, asesor_nombre, cliente_nombre, cliente_email,
        observaciones, approve_url, ai_opinion,
    )
    try:
        send_email(asesor_email, f"[REVISAR] Tasacion para {cliente_nombre}", asesor_html, attachments)
        log.info("Tasacion review email sent to asesor %s", asesor_email)
    except Exception as e:
        log.exception("Failed to email asesor %s", asesor_email)
        return jsonify({"error": f"email to asesor failed: {e}"}), 500

    return jsonify({"ok": True, "message": f"Review email sent to {asesor_email}"}), 200


@app.route("/tasacion/aprobar", methods=["GET"])
def tasacion_aprobar():
    """
    GET /tasacion/aprobar?token=xxx
    Asesor clicks this to approve and send the comparison email + Excel to the client.
    """
    token = request.args.get("token", "").strip()
    if not token or token not in _pending_tasaciones:
        return "<h2>Link invalido o expirado. Pedile al asesor que reenvie el formulario.</h2>", 400

    pending = _pending_tasaciones.pop(token)

    if time.time() - pending["created_at"] > 48 * 3600:
        return "<h2>Este link expiro (48 horas). Pedi uno nuevo desde Tally.</h2>", 400

    try:
        send_email(
            pending["cliente_email"],
            pending["subject"],
            pending["html"],
            pending.get("attachments"),
        )
        log.info("Tasacion email sent to client %s after asesor approval", pending["cliente_email"])
    except Exception as e:
        log.exception("Failed to send tasacion to client %s", pending["cliente_email"])
        return f"<h2>Error enviando el email: {e}</h2>", 500

    cliente = pending.get("cliente_nombre", "el cliente")
    asesor  = pending.get("asesor_nombre", "")
    return f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Enviado</title></head>
<body style="font-family:Arial,sans-serif;max-width:500px;margin:80px auto;text-align:center;color:#1B3A6B;">
  <img src="https://d1v2p1s05qqabi.cloudfront.net/sites/38/media/163287410792.jpeg?v=18" width="100" style="border-radius:6px;margin-bottom:24px;"><br>
  <h2 style="color:#16a34a;">&#10003; Email enviado a {cliente}</h2>
  <p style="color:#6b7280;">El informe de tasacion fue enviado correctamente. Podes cerrar esta ventana.</p>
  <p style="color:#9ca3af;font-size:12px;">Reynolds Propiedades</p>
</body></html>""", 200


@app.route("/tasacion", methods=["GET"])
def tasacion_endpoint():
    """
    GET /tasacion?barrio=Palermo&tipo=departamento&operacion=venta&superficie=85
    Returns market price stats from Argenprop for property valuation.
    """
    barrio = request.args.get("barrio", "").strip()
    if not barrio:
        return jsonify({"error": "barrio is required"}), 400

    tipo      = request.args.get("tipo", "departamento").strip()
    operacion = request.args.get("operacion", "venta").strip()
    superficie_raw = request.args.get("superficie")
    superficie = None
    if superficie_raw:
        try:
            superficie = float(superficie_raw)
        except ValueError:
            return jsonify({"error": "superficie must be a number"}), 400

    try:
        result = tasacion_module.tasacion(barrio, tipo, operacion, superficie)
        return jsonify(result), 200
    except Exception as e:
        log.exception("Error in tasacion endpoint")
        return jsonify({"error": str(e)}), 500


@app.route("/tasacion/enviar", methods=["POST"])
def tasacion_enviar():
    """
    POST /tasacion/enviar
    Body JSON: {
        "barrio": "Palermo",
        "tipo": "departamento",
        "operacion": "venta",
        "superficie": 85,
        "cliente_nombre": "Juan Perez",
        "cliente_email": "juan@example.com",
        "asesor": "Martin Reynolds"
    }
    Runs the Argenprop scraper and sends the tasacion email with market data.
    """
    body = request.get_json(force=True, silent=True) or {}

    barrio         = (body.get("barrio") or "").strip()
    cliente_nombre = (body.get("cliente_nombre") or "").strip()
    cliente_email  = (body.get("cliente_email") or "").strip()

    if not barrio or not cliente_nombre or not cliente_email:
        return jsonify({"error": "barrio, cliente_nombre and cliente_email are required"}), 400

    tipo       = (body.get("tipo") or "departamento").strip()
    operacion  = (body.get("operacion") or "venta").strip()
    asesor     = (body.get("asesor") or "tu asesor").strip()
    superficie = body.get("superficie")
    if superficie:
        try:
            superficie = float(superficie)
        except (ValueError, TypeError):
            superficie = None

    try:
        data = tasacion_module.tasacion(barrio, tipo, operacion, superficie)
    except Exception as e:
        log.exception("Tasacion scraping failed")
        return jsonify({"error": f"scraping failed: {e}"}), 500

    stats = data.get("stats")
    if not stats:
        return jsonify({"error": "No market data found for this search", "data": data}), 422

    # Build estimated value block (only if superficie provided)
    estimado_block = ""
    if data.get("estimado"):
        est = data["estimado"]
        estimado_block = f"""
            <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:28px;border-radius:10px;overflow:hidden;background:#fff7ed;border:1px solid #fed7aa;">
              <tr>
                <td style="padding:24px 28px;">
                  <p style="margin:0 0 12px;font-size:11px;font-weight:700;color:#c2410c;letter-spacing:2px;text-transform:uppercase;">Estimacion para {est['superficie_m2']:.0f} m&sup2;</p>
                  <table width="100%" cellpadding="0" cellspacing="0">
                    <tr>
                      <td style="font-size:13px;color:#555555;padding-bottom:8px;">Valor estimado por mediana</td>
                      <td align="right" style="font-size:22px;font-weight:800;color:#c2410c;padding-bottom:8px;">USD {est['valor_por_mediana']:,}</td>
                    </tr>
                    <tr>
                      <td style="font-size:12px;color:#888888;">Rango del mercado</td>
                      <td align="right" style="font-size:13px;color:#888888;">USD {est['rango_min']:,} &mdash; USD {est['rango_max']:,}</td>
                    </tr>
                  </table>
                </td>
              </tr>
            </table>"""

    # Load and fill template
    template = load_template("tasacion")
    if not template:
        return jsonify({"error": "template not found"}), 500

    html = (template
        .replace("{NOMBRE_CLIENTE}", cliente_nombre)
        .replace("{ASESOR}", asesor)
        .replace("{BARRIO}", barrio)
        .replace("{TIPO_PROPIEDAD}", tipo)
        .replace("{LISTINGS_COUNT}", str(stats["count"]))
        .replace("{USD_M2_PROMEDIO}", f"{stats['promedio_usd_m2']:,}")
        .replace("{USD_M2_MEDIANA}", f"{stats['mediana_usd_m2']:,}")
        .replace("{USD_M2_MIN}", f"{stats['minimo_usd_m2']:,}")
        .replace("{USD_M2_MAX}", f"{stats['maximo_usd_m2']:,}")
        .replace("{ESTIMADO_BLOCK}", estimado_block)
    )

    subject = f"Tasacion de tu {tipo} en {barrio} — Reynolds Propiedades"
    try:
        send_email(cliente_email, subject, html)
        log.info("Tasacion email sent to %s (%s, %s)", cliente_email, tipo, barrio)
    except Exception as e:
        log.exception("Failed to send tasacion email to %s", cliente_email)
        return jsonify({"error": f"email send failed: {e}"}), 500

    return jsonify({"ok": True, "stats": stats, "estimado": data.get("estimado")}), 200


def properties_sync_loop():
    while True:
        try:
            sync_properties.run()
        except Exception:
            log.exception("Error in properties sync")
        time.sleep(3600)


def contacts_sync_loop():
    time.sleep(60)  # wait 1 min after startup before first run
    while True:
        try:
            log.info("Starting contacts sync...")
            sync_contacts.run()
            log.info("Contacts sync complete")
        except Exception:
            log.exception("Error in contacts sync")
        time.sleep(3600)


def campaigns_loop():
    time.sleep(3600)
    while True:
        now = datetime.now()
        if now.weekday() == 6 and now.hour == 20:  # Sunday 8pm
            try:
                create_campaigns.run()
            except Exception:
                log.exception("Error creating campaigns")
        time.sleep(3600)


if __name__ == "__main__":
    threading.Thread(target=properties_sync_loop, daemon=True).start()
    threading.Thread(target=contacts_sync_loop, daemon=True).start()
    threading.Thread(target=campaigns_loop, daemon=True).start()
    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port)
